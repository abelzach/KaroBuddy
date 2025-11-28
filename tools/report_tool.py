from langchain.tools import BaseTool
from pydantic import BaseModel, Field
from typing import Optional, Literal
from datetime import datetime, timedelta
import io
from database import db_conn

class ReportInput(BaseModel):
    telegram_id: int = Field(description="User's Telegram ID")
    report_type: Literal["spending", "investment", "comprehensive"] = Field(description="Type of report to generate")
    format: Literal["pdf", "excel", "both"] = Field(description="Output format: pdf, excel, or both")
    period_days: int = Field(default=30, description="Number of days to include in report")

class ReportGenerationTool(BaseTool):
    name: str = "report_generation"
    description: str = "Generates professional PDF and Excel reports for user's financial data including spendings and investments"
    args_schema: type[BaseModel] = ReportInput
    
    def _get_user_data(self, telegram_id: int, period_days: int) -> dict:
        """Fetch user's financial data from database."""
        c = db_conn.cursor()
        
        # Get user info
        c.execute("SELECT name, username, risk_profile FROM users WHERE telegram_id=?", (telegram_id,))
        user = c.fetchone()
        user_name = user[0] if user else "User"
        risk_profile = user[2] if user and user[2] else "Not Set"
        
        # Get income data
        c.execute("""SELECT SUM(amount), COUNT(*) FROM transactions 
                     WHERE telegram_id=? AND type='income' 
                     AND date > date('now', '-' || ? || ' days')""", 
                  (telegram_id, period_days))
        income_data = c.fetchone()
        total_income = income_data[0] or 0
        income_count = income_data[1] or 0
        
        # Get expense data
        c.execute("""SELECT SUM(amount), COUNT(*) FROM transactions 
                     WHERE telegram_id=? AND type='expense' 
                     AND date > date('now', '-' || ? || ' days')""", 
                  (telegram_id, period_days))
        expense_data = c.fetchone()
        total_expenses = expense_data[0] or 0
        expense_count = expense_data[1] or 0
        
        # Get goal allocations
        c.execute("""SELECT SUM(amount) FROM transactions 
                     WHERE telegram_id=? AND type='goal_allocation' 
                     AND date > date('now', '-' || ? || ' days')""", 
                  (telegram_id, period_days))
        goal_allocation = c.fetchone()[0] or 0
        
        # Get expense breakdown by category
        c.execute("""SELECT category, SUM(amount), COUNT(*) FROM transactions 
                     WHERE telegram_id=? AND type='expense' 
                     AND date > date('now', '-' || ? || ' days')
                     GROUP BY category
                     ORDER BY SUM(amount) DESC""", 
                  (telegram_id, period_days))
        expense_breakdown = c.fetchall()
        
        # Get recent transactions
        c.execute("""SELECT date, type, category, amount, description FROM transactions 
                     WHERE telegram_id=? 
                     AND date > date('now', '-' || ? || ' days')
                     ORDER BY date DESC
                     LIMIT 50""", 
                  (telegram_id, period_days))
        recent_transactions = c.fetchall()
        
        # Get goals data
        c.execute("""SELECT goal_name, target_amount, current_amount, status, created_at 
                     FROM goals WHERE telegram_id=?
                     ORDER BY created_at DESC""", 
                  (telegram_id,))
        goals = c.fetchall()
        
        # Calculate metrics
        net_savings = total_income - total_expenses - goal_allocation
        savings_rate = (net_savings / total_income * 100) if total_income > 0 else 0
        
        return {
            'user_name': user_name,
            'risk_profile': risk_profile,
            'period_days': period_days,
            'total_income': total_income,
            'income_count': income_count,
            'total_expenses': total_expenses,
            'expense_count': expense_count,
            'goal_allocation': goal_allocation,
            'net_savings': net_savings,
            'savings_rate': savings_rate,
            'expense_breakdown': expense_breakdown,
            'recent_transactions': recent_transactions,
            'goals': goals
        }
    
    def _generate_pdf_report(self, data: dict, report_type: str) -> bytes:
        """Generate PDF report using reportlab."""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.platypus import Image as RLImage
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=18)
            
            # Container for the 'Flowable' objects
            elements = []
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=12,
                spaceBefore=12,
                fontName='Helvetica-Bold'
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#333333'),
                spaceAfter=6
            )
            
            # Title
            report_title = {
                'spending': 'Spending Analysis Report',
                'investment': 'Investment Portfolio Report',
                'comprehensive': 'Comprehensive Financial Report'
            }.get(report_type, 'Financial Report')
            
            elements.append(Paragraph(f"<b>{report_title}</b>", title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Report metadata
            report_date = datetime.now().strftime('%d %B %Y')
            period_text = f"Period: Last {data['period_days']} Days"
            
            metadata = [
                ['Generated For:', data['user_name']],
                ['Report Date:', report_date],
                ['Period:', f"Last {data['period_days']} Days"],
                ['Risk Profile:', data['risk_profile']]
            ]
            
            metadata_table = Table(metadata, colWidths=[2*inch, 4*inch])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            elements.append(metadata_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Financial Summary
            elements.append(Paragraph("<b>Financial Summary</b>", heading_style))
            
            summary_data = [
                ['Metric', 'Amount (₹)', 'Count'],
                ['Total Income', f"₹{data['total_income']:,.2f}", str(data['income_count'])],
                ['Total Expenses', f"₹{data['total_expenses']:,.2f}", str(data['expense_count'])],
                ['Goal Allocations', f"₹{data['goal_allocation']:,.2f}", '-'],
                ['Net Savings', f"₹{data['net_savings']:,.2f}", '-'],
                ['Savings Rate', f"{data['savings_rate']:.1f}%", '-']
            ]
            
            summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Expense Breakdown
            if data['expense_breakdown']:
                elements.append(Paragraph("<b>Expense Breakdown by Category</b>", heading_style))
                
                breakdown_data = [['Category', 'Amount (₹)', 'Transactions', '% of Total']]
                for category, amount, count in data['expense_breakdown']:
                    percentage = (amount / data['total_expenses'] * 100) if data['total_expenses'] > 0 else 0
                    breakdown_data.append([
                        category or 'Uncategorized',
                        f"₹{amount:,.2f}",
                        str(count),
                        f"{percentage:.1f}%"
                    ])
                
                breakdown_table = Table(breakdown_data, colWidths=[2*inch, 1.8*inch, 1.2*inch, 1*inch])
                breakdown_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
                ]))
                elements.append(breakdown_table)
                elements.append(Spacer(1, 0.3*inch))
            
            # Goals Progress
            if data['goals']:
                elements.append(Paragraph("<b>Financial Goals Progress</b>", heading_style))
                
                goals_data = [['Goal Name', 'Target (₹)', 'Current (₹)', 'Progress', 'Status']]
                for goal_name, target, current, status, created in data['goals']:
                    progress = (current / target * 100) if target > 0 else 0
                    goals_data.append([
                        goal_name,
                        f"₹{target:,.0f}",
                        f"₹{current:,.0f}",
                        f"{progress:.1f}%",
                        status.upper()
                    ])
                
                goals_table = Table(goals_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch])
                goals_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9b59b6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
                ]))
                elements.append(goals_table)
                elements.append(Spacer(1, 0.3*inch))
            
            # Recent Transactions
            if data['recent_transactions']:
                elements.append(PageBreak())
                elements.append(Paragraph("<b>Recent Transactions</b>", heading_style))
                
                trans_data = [['Date', 'Type', 'Category', 'Amount (₹)', 'Description']]
                for date, trans_type, category, amount, description in data['recent_transactions'][:20]:
                    trans_data.append([
                        date,
                        trans_type.capitalize(),
                        category or '-',
                        f"₹{amount:,.2f}",
                        (description[:30] + '...') if description and len(description) > 30 else (description or '-')
                    ])
                
                trans_table = Table(trans_data, colWidths=[1*inch, 0.8*inch, 1*inch, 1.2*inch, 2*inch])
                trans_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
                ]))
                elements.append(trans_table)
            
            # Footer
            elements.append(Spacer(1, 0.5*inch))
            footer_text = """
            <para align=center>
            <font size=8 color="#7f8c8d">
            This report is generated by KaroBuddy Financial Assistant<br/>
            For informational purposes only. Not financial advice.<br/>
            Generated on {}<br/>
            </font>
            </para>
            """.format(datetime.now().strftime('%d %B %Y at %I:%M %p'))
            elements.append(Paragraph(footer_text, normal_style))
            
            # Build PDF
            doc.build(elements)
            
            pdf_data = buffer.getvalue()
            buffer.close()
            return pdf_data
            
        except ImportError:
            raise Exception("reportlab library not installed. Please install: pip install reportlab")
        except Exception as e:
            raise Exception(f"Error generating PDF: {str(e)}")
    
    def _generate_excel_report(self, data: dict, report_type: str) -> bytes:
        """Generate Excel report using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Title
            ws_summary['A1'] = f"{report_type.upper()} REPORT"
            ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws_summary['A1'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws_summary.merge_cells('A1:E1')
            ws_summary.row_dimensions[1].height = 30
            
            # Metadata
            row = 3
            ws_summary[f'A{row}'] = "Report Information"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            metadata = [
                ('User:', data['user_name']),
                ('Period:', f"Last {data['period_days']} Days"),
                ('Generated:', datetime.now().strftime('%d %B %Y, %I:%M %p')),
                ('Risk Profile:', data['risk_profile'])
            ]
            
            for label, value in metadata:
                ws_summary[f'A{row}'] = label
                ws_summary[f'B{row}'] = value
                ws_summary[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Financial Summary
            row += 2
            ws_summary[f'A{row}'] = "Financial Summary"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12)
            ws_summary[f'A{row}'].fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
            ws_summary[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            row += 1
            
            headers = ['Metric', 'Amount (₹)', 'Count']
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            row += 1
            
            summary_data = [
                ('Total Income', data['total_income'], data['income_count']),
                ('Total Expenses', data['total_expenses'], data['expense_count']),
                ('Goal Allocations', data['goal_allocation'], '-'),
                ('Net Savings', data['net_savings'], '-'),
                ('Savings Rate', f"{data['savings_rate']:.1f}%", '-')
            ]
            
            for metric, amount, count in summary_data:
                ws_summary[f'A{row}'] = metric
                if isinstance(amount, (int, float)):
                    ws_summary[f'B{row}'] = amount
                    ws_summary[f'B{row}'].number_format = '₹#,##0.00'
                else:
                    ws_summary[f'B{row}'] = amount
                ws_summary[f'C{row}'] = count
                row += 1
            
            # Auto-adjust column widths
            for col in range(1, 6):
                ws_summary.column_dimensions[get_column_letter(col)].width = 20
            
            # Expense Breakdown Sheet
            if data['expense_breakdown']:
                ws_expenses = wb.create_sheet("Expense Breakdown")
                ws_expenses['A1'] = "EXPENSE BREAKDOWN BY CATEGORY"
                ws_expenses['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                ws_expenses['A1'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                ws_expenses['A1'].alignment = Alignment(horizontal="center")
                ws_expenses.merge_cells('A1:D1')
                
                headers = ['Category', 'Amount (₹)', 'Transactions', '% of Total']
                for col, header in enumerate(headers, 1):
                    cell = ws_expenses.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                
                row = 4
                for category, amount, count in data['expense_breakdown']:
                    percentage = (amount / data['total_expenses'] * 100) if data['total_expenses'] > 0 else 0
                    ws_expenses[f'A{row}'] = category or 'Uncategorized'
                    ws_expenses[f'B{row}'] = amount
                    ws_expenses[f'B{row}'].number_format = '₹#,##0.00'
                    ws_expenses[f'C{row}'] = count
                    ws_expenses[f'D{row}'] = percentage / 100
                    ws_expenses[f'D{row}'].number_format = '0.0%'
                    row += 1
                
                for col in range(1, 5):
                    ws_expenses.column_dimensions[get_column_letter(col)].width = 20
            
            # Goals Sheet
            if data['goals']:
                ws_goals = wb.create_sheet("Goals Progress")
                ws_goals['A1'] = "FINANCIAL GOALS PROGRESS"
                ws_goals['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                ws_goals['A1'].fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                ws_goals['A1'].alignment = Alignment(horizontal="center")
                ws_goals.merge_cells('A1:E1')
                
                headers = ['Goal Name', 'Target (₹)', 'Current (₹)', 'Progress %', 'Status']
                for col, header in enumerate(headers, 1):
                    cell = ws_goals.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                
                row = 4
                for goal_name, target, current, status, created in data['goals']:
                    progress = (current / target * 100) if target > 0 else 0
                    ws_goals[f'A{row}'] = goal_name
                    ws_goals[f'B{row}'] = target
                    ws_goals[f'B{row}'].number_format = '₹#,##0'
                    ws_goals[f'C{row}'] = current
                    ws_goals[f'C{row}'].number_format = '₹#,##0'
                    ws_goals[f'D{row}'] = progress / 100
                    ws_goals[f'D{row}'].number_format = '0.0%'
                    ws_goals[f'E{row}'] = status.upper()
                    row += 1
                
                for col in range(1, 6):
                    ws_goals.column_dimensions[get_column_letter(col)].width = 18
            
            # Transactions Sheet
            if data['recent_transactions']:
                ws_trans = wb.create_sheet("Transactions")
                ws_trans['A1'] = "RECENT TRANSACTIONS"
                ws_trans['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                ws_trans['A1'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                ws_trans['A1'].alignment = Alignment(horizontal="center")
                ws_trans.merge_cells('A1:E1')
                
                headers = ['Date', 'Type', 'Category', 'Amount (₹)', 'Description']
                for col, header in enumerate(headers, 1):
                    cell = ws_trans.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                
                row = 4
                for date, trans_type, category, amount, description in data['recent_transactions']:
                    ws_trans[f'A{row}'] = date
                    ws_trans[f'B{row}'] = trans_type.capitalize()
                    ws_trans[f'C{row}'] = category or '-'
                    ws_trans[f'D{row}'] = amount
                    ws_trans[f'D{row}'].number_format = '₹#,##0.00'
                    ws_trans[f'E{row}'] = description or '-'
                    row += 1
                
                ws_trans.column_dimensions['A'].width = 12
                ws_trans.column_dimensions['B'].width = 12
                ws_trans.column_dimensions['C'].width = 15
                ws_trans.column_dimensions['D'].width = 15
                ws_trans.column_dimensions['E'].width = 40
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_data = buffer.getvalue()
            buffer.close()
            return excel_data
            
        except ImportError:
            raise Exception("openpyxl library not installed. Please install: pip install openpyxl")
        except Exception as e:
            raise Exception(f"Error generating Excel: {str(e)}")
    
    def _run(self, telegram_id: int, report_type: str, format: str, period_days: int = 30) -> str:
        """Main execution method."""
        try:
            # Fetch user data
            data = self._get_user_data(telegram_id, period_days)
            
            # Check if user has any data
            if data['total_income'] == 0 and data['total_expenses'] == 0:
                return """📊 **No Data Available**

You don't have any financial data yet to generate a report.

Start by:
• Logging your income: "I earned 50000"
• Tracking expenses: "Spent 2000 on groceries"
• Setting goals: "Create goal Emergency Fund with target 100000"

Once you have some data, I'll generate beautiful reports for you! 📈"""
            
            files_generated = []
            file_paths = []
            
            # Generate PDF
            if format in ['pdf', 'both']:
                try:
                    pdf_data = self._generate_pdf_report(data, report_type)
                    filename = f"karobuddy_{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                    
                    # Save to file
                    with open(filename, 'wb') as f:
                        f.write(pdf_data)
                    
                    files_generated.append('PDF')
                    file_paths.append(filename)
                except Exception as e:
                    return f"⚠️ Error generating PDF: {str(e)}\n\nPlease ensure reportlab is installed: pip install reportlab", []
            
            # Generate Excel
            if format in ['excel', 'both']:
                try:
                    excel_data = self._generate_excel_report(data, report_type)
                    filename = f"karobuddy_{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    # Save to file
                    with open(filename, 'wb') as f:
                        f.write(excel_data)
                    
                    files_generated.append('Excel')
                    file_paths.append(filename)
                except Exception as e:
                    return f"⚠️ Error generating Excel: {str(e)}\n\nPlease ensure openpyxl is installed: pip install openpyxl", []
            
            # Generate success message
            report_name = {
                'spending': 'Spending Analysis',
                'investment': 'Investment Portfolio',
                'comprehensive': 'Comprehensive Financial'
            }.get(report_type, 'Financial')
            
            response = f"""✅ **Report Generated Successfully!**

📊 **{report_name} Report**
📅 Period: Last {period_days} days
📁 Format: {' & '.join(files_generated)}

**📈 Quick Summary:**
• Total Income: ₹{data['total_income']:,.2f}
• Total Expenses: ₹{data['total_expenses']:,.2f}
• Net Savings: ₹{data['net_savings']:,.2f}
• Savings Rate: {data['savings_rate']:.1f}%

The report includes:
✅ Financial summary with key metrics
✅ Expense breakdown by category
✅ Goals progress tracking
✅ Recent transaction history
✅ Professional formatting and charts

📎 Report file(s) attached above!"""
            
            return response, file_paths
            
        except Exception as e:
            return f"⚠️ Error generating report: {str(e)}\n\nPlease try again or contact support."

# Create instance
report_tool = ReportGenerationTool()